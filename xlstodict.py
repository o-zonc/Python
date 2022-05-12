import openpyxl as o

wb = o.load_workbook("D:\Asterisk\신촌캠 노드.xlsx")
ws = wb['Sheet1']

style = {}
for i in range(2, 166):
    name = ws.cell(row = i, column = 1).value
    com = [name, 10000]
    help = []
    for j in range(4, 14, 2):
        list = []
        temp1 = ws.cell(row = i, column = j).value
        temp2 = ws.cell(row = i, column = j + 1).value
        if temp1 == None:
            pass
        else:
            list.append(temp1)
            list.append(temp2)
            help.append(list)
    com.append(help)
    style[name] = com

print(style)