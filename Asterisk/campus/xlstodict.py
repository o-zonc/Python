import openpyxl as o

wb = o.load_workbook("D:\Asterisk\신촌캠 노드.xlsx")
ws = wb['Sheet1']

node = {}
for i in range(2, 167):
    nodename = ws.cell(row = i, column = 1).value
    tmplt = [nodename, 10000]
    guide = []
    for j in range(4, 20, 2):
        list = []
        nodecode = ws.cell(row = i, column = j).value
        cloud = str(ws.cell(row = i, column = j + 1).value)
        cloud = cloud.replace(" ", "")
        cloud = cloud.split(',')
        if nodecode == None or cloud == None:
            pass
        else:
            list.append(nodecode)
            list.append(int(cloud[0]))
            list.append(int(cloud[1]))
            guide.append(list)
    tmplt.append(guide)
    node[nodename] = tmplt

print(node)
get = str(node)
file = open("D:\Asterisk\campus\sinchon.py",'w')
file.write(get)
file.close()